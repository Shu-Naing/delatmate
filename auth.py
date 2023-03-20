from flask import Blueprint, render_template, redirect, url_for, request, flash, Response, current_app as app
from flask import Flask, request, jsonify
from werkzeug.security import generate_password_hash, check_password_hash
from models import Mocdm_users,Mocdm_erp,Mocdm_pending,Mocdm_consumption,Mocdm_schedule
from os.path import join, dirname, realpath
from flask import Flask, session
from flask_login import login_user, logout_user, login_required, current_user
from flask_security import roles_accepted
from io import TextIOWrapper
from flask import make_response
from sqlalchemy import exc, func
from sqlalchemy.exc import DataError, IntegrityError
from sqlalchemy.exc import SQLAlchemyError
from datetime import datetime
from datetime import timedelta
from flask import Flask, send_file
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker
from flask import current_app
from jinja2 import Environment
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from openpyxl.styles import Font
from io import BytesIO
import collections
import pyttsx3 
import csv
import logging
import traceback
import sys
import os
import pandas as pd
import xlsxwriter
import io
import re
from __init__ import db

auth = Blueprint('auth', __name__)


def is_active(page):
    return True if request.path == page else False

@auth.route('/login', methods=['GET', 'POST']) 
def login():
    try:
        if request.method=='GET': 
            return render_template('login.html')
        else: 
            name = request.form.get('name')
            password = request.form.get('password')
            remember = True if request.form.get('remember') else False
            user = Mocdm_users.query.filter_by(name=name).first()
            if not user:
                flash('Please login agin!')
                return redirect(url_for('auth.login'))
            elif not check_password_hash(user.password, password):
                flash('Please check your login details and try again.')
                return redirect(url_for('auth.login')) 
            login_user(user, remember=remember)
            session["role"] = user.role
            session["user_session"] = user.id
            current_page = current_app.config.get('CURRENT_PAGE')
            current_page_home = 'home'
            return redirect(url_for('main.profile'))
    except SQLAlchemyError as e:
        current_time = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
        logging.basicConfig(filename= f'error_log.log', level=logging.ERROR)
        logging.error(str(e))
    return redirect(url_for('main.profile')) 

@auth.route('/logout') 
@login_required
def logout(): 
    logout_user()
    return redirect(url_for('main.index'))
    
@auth.route('/change_password', methods=['GET', 'POST'])
@login_required
def change_password():
    try:
        if request.method == 'POST':
            id = request.form['id']
            current_password = request.form['current_password']
            new_password = request.form['new_password']
            
            user = Mocdm_users.query.get(id)        
            if check_password_hash(user.password, current_password):
                hashed_password = generate_password_hash(new_password)
                user.password = hashed_password
                db.session.commit()
                return redirect(url_for('auth.emplist'))
            else:
                flash('Incorrect password')
        
        return redirect(url_for('auth.emplist'))
    except SQLAlchemyError as e:
        current_time = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
        logging.basicConfig(filename= f'error_log.log', level=logging.ERROR)
        logging.error(str(e))
    return redirect(url_for('auth.emplist')) 

@auth.route('/signup', methods=['GET', 'POST'])
@login_required
def signup():
    try:
        if request.method=='GET': 
            return render_template('signup.html')
        else:
            email = request.form.get('email')
            name = request.form.get('name')
            phone = request.form.get('phone')
            password = request.form.get('password')
            role = request.form.get('role')
            user = Mocdm_users.query.filter_by(email=email).first()
            if user: 
                flash('Email address already exists')
                return redirect(url_for('auth.signup'))
            new_user = Mocdm_users(email=email, name=name,phone=phone, password=generate_password_hash(password, method='sha256'),role=role) 
            db.session.add(new_user)
            db.session.commit()
            return redirect(url_for('auth.login'))
    except SQLAlchemyError as e:
        current_time = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
        logging.basicConfig(filename= f'error_log.log', level=logging.ERROR)
        logging.error(str(e))
    return redirect(url_for('auth.erplist')) 

@auth.route('/updateEmp', methods=['GET', 'POST'])
@login_required
def updateEmp():
    try:
        if request.method == 'POST':
            got_data = Mocdm_users.query.get(request.form.get('id'))
            got_data.name = request.form['name']
            got_data.email = request.form['email']
            got_data.phone = request.form['phone']
            db.session.commit()
            return redirect(url_for('auth.emplist'))
    except SQLAlchemyError as e:
        current_time = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
        logging.basicConfig(filename= f'error_log.log', level=logging.ERROR)
        logging.error(str(e))
    return redirect(url_for('auth.emplist'))  

@auth.route('/emplist', defaults={'page_num':1})
@auth.route('/emplist/<int:page_num>')
@login_required
def emplist(page_num):
    try:
        if request.method=='GET':
            all_data = db.session.query(Mocdm_users).filter(~Mocdm_users.name.like('mocdev')).paginate(per_page=5, page=page_num, error_out=True)
    except SQLAlchemyError as e:
        current_time = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
        logging.basicConfig(filename= f'error_log.log', level=logging.ERROR)
        logging.error(str(e))
    return render_template('employee.html',all_data = all_data, emp_active="is_active('/emp')") 

@auth.route("/erpupload", methods=['POST'])
@login_required
def erpupload():
    try:
        if request.method == 'POST':
            excel_file = request.files['file']
            col_names = ['PO','款号','款号版本', '产品编号', '产品名称' , '产品主色','产品配色','货期','配色项目','物料分类','物料编号','物料英文名称','物料名称','规格','颜色','库存单位用量','库存单位','生产损耗','生产用量','订单数量','需求总用量','备注']
            df = pd.read_excel(excel_file,names=col_names,header = None,skiprows=1)
            df = df.fillna('')
            for i,row in df.iterrows():
                txt = row['产品编号']
                x = txt.split('-')
                if len(x) < 4:
                    y = x[1]
                else:
                    y = x[1]+x[2]
                c = row['款号']
                d = c.strip('')
                e = row['产品主色'] 
                f = e.upper()
                all_data = Mocdm_erp(po=row['PO'],style=d,buyer_version=row['款号版本'],buyer=row['产品编号'],product_name=row['产品名称'],main_color=f,season=row['产品配色'],vessel_date=row['货期'],category=row['配色项目'],material_classification=row['物料分类'],material_code=row['物料编号'],material=row['物料英文名称'],material_chinese=row['物料名称'],size=row['规格'],color=row['颜色'],org_consume=row['库存单位用量'],unit=row['库存单位'],loss=row['生产损耗'],consume_point=row['生产用量'],order_qty=row['订单数量'],consume=row['需求总用量'],gp=row['备注'],pending_buyer=y)
                db.session.add(all_data)
                db.session.commit()
            return redirect(url_for('auth.erplist'))
    except SQLAlchemyError as e:
        current_time = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
        logging.basicConfig(filename= f'error_log.log', level=logging.ERROR)
        logging.error(str(e))
    return redirect(url_for('auth.erplist'))
    
@auth.route('/erplist', defaults={'page_num':1}, methods=['GET', 'POST'])
@auth.route('/erplist/<int:page_num>', methods=['GET', 'POST'])
@login_required
def erplist(page_num):
    try:
        if request.method=='POST' or request.args.get('search')=='True' :
            if (request.args.get('po')): 
                po = request.args.get('po')
            else:
                po = request.form.get('po', '')

            if (request.args.get('style')): 
                style = request.args.get('style')
            else:
                style = request.form.get('style', '')

            if (request.args.get('buyer')): 
                buyer = request.args.get('buyer')
            else:
                buyer = request.form.get('buyer', '')

            if (request.args.get('main_color')): 
                main_color = request.args.get('main_color')
            else:
                main_color = request.form.get('main_color', '')

            if (request.args.get('gp')): 
                gp = request.args.get('gp')
            else:
                gp = request.form.get('gp', '')
            session['po'] = po
            session['style'] = style
            session['buyer'] = buyer
            session['main_color'] = main_color
            session['gp'] = gp
            search1 = "%{}%".format(po)
            search2 = "%{}%".format(style)
            search3 = "%{}%".format(buyer)
            search4 = "%{}%".format(main_color)
            search5 = "%{}%".format(gp)
            all_data = Mocdm_erp.query.filter((Mocdm_erp.po.like(search1)),(Mocdm_erp.style.like(search2)),(Mocdm_erp.buyer.like(search3)),(Mocdm_erp.main_color.like(search4)),(Mocdm_erp.gp.like(search5))).paginate(per_page=100, page=page_num, error_out=True)
            return render_template("erp.html",po = po,style=style,gp=gp,buyer=buyer,main_color=main_color,all_data = all_data, erp_active="is_active('/erp')")
        else:         
            all_data = Mocdm_erp.query.paginate(per_page=100, page=page_num, error_out=True)
            return render_template('erp.html',all_data = all_data, erp_active="is_active('/erp')")
    except SQLAlchemyError as e:
        current_time = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
        logging.basicConfig(filename= f'error_log.log', level=logging.ERROR)
        logging.error(str(e))
    return render_template('erp.html',po = po,gp=gp,style=style,buyer=buyer,main_color=main_color,all_data = all_data, erp_active="is_active('/erp')")
    

@auth.route('/download_erp', methods=['GET', 'POST'])
def download_erp():
    if request.args.get('search') == 'true':
        po = request.args.get('po')
        style = request.args.get('style')
        buyer = request.args.get('buyer')
        main_color = request.args.get('main_color')
        gp = request.args.get('gp')
        po = "%{}%".format(po)
        style = "%{}%".format(style)
        buyer = "%{}%".format(buyer)
        main_color = "%{}%".format(main_color)
        gp = "%{}%".format(gp)
        all_data = Mocdm_erp.query.filter((Mocdm_erp.po.like(po)),(Mocdm_erp.style.like(style)),(Mocdm_erp.buyer.like(buyer)),(Mocdm_erp.color.like(main_color)),(Mocdm_erp.gp.like(gp))).all()
        wb = Workbook()
        ws = wb.active
        ws.append(['PO','STYLE','BUYER VERSION','BUYER','PRODUCT NAME','MAIN COLOR','SEASON','VESSEL DATE','CATEGORY','MATERIAL CLASSIFICATION','MATERIAL CODE','MATERIAL' ,'MATERIAL NAME IN CHINESE','SIZE','COLOUR','ORIGINAL CONSUME','UNIT','LOSS','CONSUME POINT','ORDER QTY','CONSUME','GROUP'])
        for item in all_data:
            ws.append([item.po,item.style,item.buyer_version,item.buyer,item.product_name,item.main_color,item.season,item.vessel_date.strftime('%m/%d/%Y'),item.category,item.material_classification,item.material_code,item.material,item.material_chinese,item.size,item.color,item.org_consume,item.unit,item.loss,item.consume_point,item.order_qty,item.consume,item.gp])
        file = BytesIO()
        wb.save(file)
        file.seek(0)
        filename = 'erp.xlsx'
        r = Response(
            file.read(),
            headers={
                'Content-Disposition': f'attachment;filename={filename}',
                'Content-Type': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            }
        )
        return r
    else:
        all_data1 = Mocdm_erp.query.all()
        wb = Workbook()
        ws = wb.active
        ws.append(['PO','STYLE','BUYER VERSION','BUYER','PRODUCT NAME','MAIN COLOR','SEASON','VESSEL DATE','CATEGORY','MATERIAL CLASSIFICATION','MATERIAL CODE','MATERIAL' ,'MATERIAL NAME IN CHINESE','SIZE','COLOUR','ORIGINAL CONSUME','UNIT','LOSS','CONSUME POINT','ORDER QTY','CONSUME','GROUP'])
        for item in all_data1:
            ws.append([item.po,item.style,item.buyer_version,item.buyer,item.product_name,item.main_color,item.season,item.vessel_date.strftime('%m/%d/%Y'),item.category,item.material_classification,item.material_code,item.material,item.material_chinese,item.size,item.color,item.org_consume,item.unit,item.loss,item.consume_point,item.order_qty,item.consume,item.gp])
        file = BytesIO()
        wb.save(file)
        file.seek(0)
        filename = 'erp.xlsx'
        r = Response(
            file.read(),
            headers={
                'Content-Disposition': f'attachment;filename={filename}',
                'Content-Type': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            }
        )
        return r
        

@auth.route('/erpUpdate', methods=['GET', 'POST'])
@login_required
def erpUpdate():
    try:
        if request.method=='POST':
            all_data = Mocdm_erp.query.get(request.form.get('id'))
            all_data.po = request.form['po']
            all_data.style = request.form['style']
            all_data.buyer = request.form['buyer']
            all_data.vessel_date = request.form['vessel_date']
            all_data.season = request.form['season']
            all_data.gp = request.form['gp']
            all_data.buyer_version = request.form['buyer_version']
            all_data.product_name = request.form['product_name']
            all_data.main_color = request.form['main_color']
            all_data.category = request.form['category']
            all_data.material_classification = request.form['material_classification']
            all_data.material_code = request.form['material_code']
            all_data.material = request.form['material']
            all_data.material_chinese = request.form['material_chinese']
            all_data.size = request.form['size']
            all_data.org_consume = request.form['org_consume']
            all_data.unit = request.form['unit']
            all_data.loss = request.form['loss']
            all_data.consume_point = request.form['consume_point']
            all_data.order_qty = request.form['order_qty']
            all_data.consume = request.form['consume']
            all_data.remark = request.form['remark']
            all_data.status = request.form['status']
            db.session.commit()
    except SQLAlchemyError as e:
        current_time = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
        logging.basicConfig(filename= f'error_log.log', level=logging.ERROR)
        logging.error(str(e))
    return f"<td>{all_data.po}</td>,<td>{all_data.style}</td>,<td>{all_data.buyer_version}</td>,<td>{all_data.buyer}</td>,<td>{all_data.product_name}</td>,<td>{all_data.main_color}</td>,<td>{all_data.season}</td>,<td>{all_data.vessel_date}</td>,<td>{all_data.category}</td>,<td>{all_data.material_classification}</td>,<td>{all_data.material_code}</td>,<td>{all_data.material}</td>,<td>{all_data.material_chinese}</td>,<td>{all_data.size}</td>,<td>{all_data.color}</td>,<td>{all_data.org_consume}</td>,<td>{all_data.unit}</td>,<td>{all_data.loss}</td>,<td>{all_data.consume_point}</td>,<td>{all_data.order_qty}</td>,<td>{all_data.consume}</td>,<td>{all_data.gp}</td>,<td>{all_data.remark}</td>,<td>{all_data.status}</td>,<td><a href='/erpUpdate'  class='btn btn-primary' value='{all_data.id}' edit-row-value='{all_data.id}' data-bs-toggle='modal' data-bs-target='#myModal{all_data.id}'>Edit</a></td>"

@auth.route('/pendinglist', defaults={'page_num':1}, methods=['GET', 'POST'])
@auth.route('/pendinglist/<int:page_num>', methods=['GET','POST'])
@login_required
def pendinglist(page_num):
    if request.method=='POST' or request.args.get('search')=='True' :
            if (request.args.get('pen_po')): 
                pen_po = request.args.get('po')
            else:
                pen_po = request.form.get('pen_po', '')

            if (request.args.get('pen_style')): 
                pen_style = request.args.get('pen_style')
            else:
                pen_style = request.form.get('pen_style', '')

            if (request.args.get('pen_org_buyer')): 
                pen_org_buyer = request.args.get('pen_org_buyer')
            else:
                pen_org_buyer = request.form.get('pen_org_buyer', '')

            if (request.args.get('pen_color')): 
                pen_color = request.args.get('pen_color')
            else:
                pen_color = request.form.get('pen_color', '')

            if (request.args.get('pen_gp_name')): 
                pen_gp_name = request.args.get('pen_gp_name')
            else:
                pen_gp_name = request.form.get('pen_gp_name', '')

            if (request.args.get('pen_ext_dely')): 
                pen_ext_dely = request.args.get('pen_ext_dely')
            else:
                pen_ext_dely = request.form.get('pen_ext_dely', '')

            if (request.args.get('pen_order_date')): 
                pen_order_date = request.args.get('pen_order_date')
            else:
                pen_order_date = request.form.get('pen_order_date', '')

            if (request.args.get('pen_order_date')): 
                pen_order_date = request.args.get('pen_order_date')
            else:
                pen_order_date = request.form.get('pen_order_date', '') 
            session['pen_po'] = pen_po
            session['pen_style'] = pen_style
            session['pen_org_buyer'] = pen_org_buyer
            session['pen_color'] = pen_color
            session['pen_gp_name'] = pen_gp_name
            session['pen_ext_dely'] = pen_ext_dely
            session['pen_order_date'] = pen_order_date
            search1 = "%{}%".format(pen_po)
            search2 = "%{}%".format(pen_style)
            search3 = "%{}%".format(pen_org_buyer)
            search4 = "%{}%".format(pen_color)
            search5 = "%{}%".format(pen_gp_name)
            search6 = "%{}%".format(pen_ext_dely)
            search7 = "%{}%".format(pen_order_date)
            all_data = Mocdm_pending.query.filter((Mocdm_pending.po.like(search1)),(Mocdm_pending.style.like(search2)),(Mocdm_pending.org_buyer.like(search3)),(Mocdm_pending.color.like(search4)),(Mocdm_pending.gp_name.like(search5)),(Mocdm_pending.ext_dely.like(search6)),(Mocdm_pending.order_date.like(search7))).paginate(per_page=100, page=page_num, error_out=True)
            return render_template("pending.html",pen_pen_po = pen_po,pen_style=pen_style,pen_org_buyer=pen_org_buyer,pen_color=pen_color,pen_gp_name=pen_gp_name,pen_ext_dely=pen_ext_dely,pen_order_date=pen_order_date, all_data = all_data, pending_active="is_active('/pending')")       
    else:
        all_data = Mocdm_pending.query.paginate(per_page=100, page=page_num, error_out=True)
        return render_template('pending.html',all_data = all_data, pending_active="is_active('/pending')")


@auth.route('/download_pending', methods=['GET', 'POST'])
def download_pending():
    if request.args.get('search') == 'true':
        po = request.args.get('po')
        style = request.args.get('style')
        org_buyer = request.args.get('org_buyer')
        color = request.args.get('color')
        gp_name = request.args.get('gp_name')
        ext_dely = request.args.get('ext_dely')
        order_date = request.args.get('order_date')
        po = "%{}%".format(po)
        style = "%{}%".format(style)
        org_buyer = "%{}%".format(org_buyer)
        color = "%{}%".format(color)
        gp_name = "%{}%".format(gp_name)
        ext_dely = "%{}%".format(ext_dely)
        order_date = "%{}%".format(order_date)
        all_data = Mocdm_pending.query.filter((Mocdm_pending.po.like(po)),(Mocdm_pending.style.like(style)),(Mocdm_pending.org_buyer.like(org_buyer)),(Mocdm_pending.color.like(color)),(Mocdm_pending.gp_name.like(gp_name)),(Mocdm_pending.ext_dely.like(ext_dely)),(Mocdm_pending.order_date.like(order_date))).all()
        wb = Workbook()
        ws = wb.active
        ws.append(['EX - FTY ( DELY)','MCN ( D/C )','PO#','MYANMAR','Ship To','LABEL','Linked Store','DES','GROUP NAME','Style#','Buyer#','COLOUR','QTY','Vessel','Factory Name','DB/GB Pkg Code','SDN PO','Customer Po#','UPC Number','Linked SO Num',"Ref.Number",'Material Lot No:','Season','Buyer','ORDER DATE','KZM ID','Remark','ShpgJob','xFty Date','Status'])
        for item in all_data:
            if item.xfty_date == None or '':
                x = item.xfty_date
            else:
                x = item.xfty_date.strftime('%m/%d/%Y')
            ws.append([item.ext_dely.strftime('%m/%d/%Y'),item.mcn,item.po,item.myanmar,item.ship_to,item.label,item.linked_store,item.des,item.gp_name,item.style,item.org_buyer,item.color,item.qty,item.vessel_date.strftime('%d/%m/%Y'),item.factory,item.db_gb_code,item.sdn_po,item.customer_po,item.upc_no,item.linked_so_no,item.ref_no,item.material_log_no,item.season,item.buyer_txt,item.order_date.strftime('%d/%m/%Y'),item.kmz_id,item.remark,item.shpg_job,x,item.status])
        file = BytesIO()
        wb.save(file)
        file.seek(0)
        filename = 'pending.xlsx'
        r = Response(
            file.read(),
            headers={
                'Content-Disposition': f'attachment;filename={filename}',
                'Content-Type': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            }
        )
        return r
    else:
        all_data1 = Mocdm_pending.query.all()
        wb = Workbook()
        ws = wb.active
        ws.append(['EX - FTY ( DELY)','MCN ( D/C )','PO#','MYANMAR','Ship To','LABEL','Linked Store','DES','GROUP NAME','Style#','Buyer#','COLOUR','QTY','Vessel','Factory Name','DB/GB Pkg Code','SDN PO','Customer Po#','UPC Number','Linked SO Num',"Ref.Number",'Material Lot No:','Season','Buyer','ORDER DATE','KZM ID','Remark','ShpgJob','xFty Date','Status'])
        for item in all_data1:
            if item.xfty_date == None or '':
                x = item.xfty_date
            else:
                x = item.xfty_date.strftime('%m/%d/%Y')
            ws.append([item.ext_dely.strftime('%m/%d/%Y'),item.mcn,item.po,item.myanmar,item.ship_to,item.label,item.linked_store,item.des,item.gp_name,item.style,item.org_buyer,item.color,item.qty,item.vessel_date.strftime('%d/%m/%Y'),item.factory,item.db_gb_code,item.sdn_po,item.customer_po,item.upc_no,item.linked_so_no,item.ref_no,item.material_log_no,item.season,item.buyer_txt,item.order_date.strftime('%d/%m/%Y'),item.kmz_id,item.remark,item.shpg_job,x,item.status])
        file = BytesIO()
        wb.save(file)
        file.seek(0)
        filename = 'pending.xlsx'
        r = Response(
            file.read(),
            headers={
                'Content-Disposition': f'attachment;filename={filename}',
                'Content-Type': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            }
        )
        return r

@auth.route('/pendingUpdate', methods=['GET', 'POST'])
@login_required
def pendingUpdate():
    try:
        if request.method=='POST':
            all_data = Mocdm_pending.query.get(request.form.get('id'))
            all_data.ext_dely = request.form['ext_dely']
            all_data.po = request.form['po']
            all_data.myanmar = request.form['myanmar']
            all_data.gp_name = request.form['gp_name']
            all_data.style = request.form['style']
            all_data.color = request.form['color']
            all_data.vessel_date = request.form['vessel_date']
            all_data.customer_po = request.form['customer_po']
            all_data.buyer_txt = request.form['buyer_txt']
            all_data.order_date = request.form['order_date']
            if not request.form['xfty_date']:
                l = None
            else:
                l = datetime.strptime(request.form['xfty_date'], '%Y-%m-%d')
            all_data.xfty_date = l
            all_data.mcn = request.form['mcn']
            all_data.ship_to = request.form['ship_to']
            all_data.label = request.form['label']
            all_data.linked_store = request.form['linked_store']
            all_data.des = request.form['des']
            all_data.qty = request.form['qty']
            all_data.factory = request.form['factory']
            all_data.db_gb_code = request.form['db_gb_code']
            all_data.upc_no = request.form['upc_no']
            all_data.linked_so_no = request.form['linked_so_no']
            all_data.ref_no = request.form['linked_so_no']
            all_data.linked_so_no = request.form['ref_no']
            all_data.material_log_no = request.form['material_log_no']
            all_data.season = request.form['season']
            all_data.kmz_id = request.form['kmz_id']
            all_data.remark = request.form['remark']
            all_data.shpg_job = request.form['shpg_job']
            all_data.status = request.form['status']
            db.session.commit()
            if not all_data.xfty_date:
                w = all_data.xfty_date
            else:
                w = all_data.xfty_date.strftime('%m/%d/%Y')
            
    except SQLAlchemyError as e:
        current_time = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
        logging.basicConfig(filename= f'error_log.log', level=logging.ERROR)
        logging.error(str(e))
    # return redirect(url_for('auth.pendinglist'))
    return f"<td>{all_data.ext_dely}</td>,<td>{all_data.mcn}</td>,<td>{all_data.po}</td>,<td>{all_data.myanmar}</td>,<td>{all_data.ship_to}</td>,<td>{all_data.label}</td>,<td>{all_data.linked_store}</td>,<td>{all_data.des}</td>,<td>{all_data.gp_name}</td>,<td>{all_data.style}</td>,<td>{all_data.org_buyer}</td>,<td>{all_data.color}</td>,<td>{all_data.qty}</td>,<td>{all_data.vessel_date}</td>,<td>{all_data.factory}</td>,<td>{all_data.db_gb_code}</td>,<td>{all_data.sdn_po}</td>,<td>{all_data.customer_po}</td>,<td>{all_data.upc_no}</td>,<td>{all_data.linked_so_no}</td>,<td>{all_data.ref_no}</td>,<td>{all_data.material_log_no}</td>,<td>{all_data.season}</td>,<td>{all_data.buyer_txt}</td>,<td>{all_data.order_date}</td>,<td>{all_data.kmz_id}</td>,<td>{all_data.remark}</td>,<td>{all_data.shpg_job}</td>,<td>{w}</td>,<td>{all_data.status}</td>,<td><a href='/pendingUpdate'  class='btn btn-primary' value='{all_data.id}' edit-row-value='{all_data.id}' data-bs-toggle='modal' data-bs-target='#myModal{all_data.id}'>Edit</a></td>"

@auth.route("/pending_upload", methods=['POST'])
@login_required
def pending_upload():
    try:
        if request.method == 'POST':
            excel_file = request.files['file']
            col_names = ['Ex-Fty','MCN','PO#', 'Ship To', 'Label' , 'Linked Store','DES','Group Name','Style#','Buyer#','COLOUR','QTY','Vessel','Factory','DB/GB Pkg Code','SDN PO','Customer Po#','UPC Number','Linked SO Num','Ref.Number','Material Lot No:','Season','Buyer','Order Date','KZM ID','Remark','ShpgJob','xFty Date']
            df = pd.read_excel(excel_file,names=col_names,header = None,skiprows=2)
            df = df.fillna('')
            for i,row in df.iterrows():
                txt = 'MYANMAR'
                x = row['Buyer#']
                h = x[:11]
                c = row['Style#']
                d = c.replace(" ", "")
                e = row['COLOUR']
                f = e.upper()
                trimmed_text = re.sub(r'^\s+|\s+$', '', f)
                if not row['xFty Date']:
                    l = None
                else:
                    l = datetime.strptime(j, '%Y-%m-%d')
                all_data = Mocdm_pending(ext_dely=row['Ex-Fty'],mcn=row['MCN'],po=row['PO#'],ship_to=row['Ship To'],label=row['Label'],linked_store=row['Linked Store'],des=row['DES'],gp_name=row['Group Name'],style=d,org_buyer=h,color=trimmed_text,qty=row['QTY'],vessel_date=row['Vessel'],factory=row['Factory'],db_gb_code=row['DB/GB Pkg Code'],sdn_po=row['SDN PO'],customer_po=row['Customer Po#'],upc_no=row['UPC Number'],linked_so_no=row['Linked SO Num'],ref_no=row['Ref.Number'],material_log_no=row['Material Lot No:'],season=row['Season'],buyer_txt=row['Buyer'],order_date=row['Order Date'],kmz_id=row['KZM ID'],remark=row['Remark'],shpg_job=row['ShpgJob'],xfty_date=l,myanmar=txt,previous=row['Ex-Fty'])
                db.session.add(all_data)
                db.session.commit()
            return redirect(url_for('auth.pendinglist'))
    except SQLAlchemyError as e:
        current_time = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
        logging.basicConfig(filename= f'error_log.log', level=logging.ERROR)
        logging.error(str(e))
    return redirect(url_for('auth.pendinglist'))

@auth.route('/deletePending', methods=['POST'])
def deletePending():
    start_date = request.form['start_date']
    end_date = request.form['end_date']
    
    start_datetime = datetime.strptime(start_date, '%Y-%m-%d')
    end_datetime = datetime.strptime(end_date, '%Y-%m-%d')
    
    all_data = Mocdm_pending.query.filter(Mocdm_pending.ext_dely >= start_datetime, Mocdm_pending.ext_dely <= end_datetime).delete()
    db.session.commit()
    return redirect(url_for('auth.pendinglist'))

@auth.route('/orderlist', defaults={'page_num':1}, methods=['GET','POST'])
@auth.route('/orderlist/<int:page_num>', methods=['GET','POST'])
@login_required
def orderlist(page_num):
    try:
        if request.method=='POST' or request.args.get('search')=='True' :
            if (request.args.get('or_po')): 
                or_po = request.args.get('or_po')
            else:
                or_po = request.form.get('or_po', '')

            if (request.args.get('or_style')): 
                or_style = request.args.get('or_style')
            else:
                or_style = request.form.get('or_style', '')

            if (request.args.get('or_org_buyer')): 
                or_org_buyer = request.args.get('or_org_buyer')
            else:
                or_org_buyer = request.form.get('or_org_buyer', '')

            if (request.args.get('or_color')): 
                or_color = request.args.get('or_color')
            else:
                or_color = request.form.get('or_color', '')

            if (request.args.get('or_gp_name')): 
                or_gp_name = request.args.get('or_gp_name')
            else:
                or_gp_name = request.form.get('or_gp_name', '')

            if (request.args.get('or_ext_dely')): 
                or_ext_dely = request.args.get('or_ext_dely')
            else:
                or_ext_dely = request.form.get('or_ext_dely', '')

            if (request.args.get('or_order_date')): 
                or_order_date = request.args.get('or_order_date')
            else:
                or_order_date = request.form.get('or_order_date', '')

            if (request.args.get('or_factory')): 
                or_factory = request.args.get('or_factory')
            else:
                or_factory = request.form.get('or_factory', '')

            if (request.args.get('or_label')): 
                or_label = request.args.get('or_label')
            else:
                or_label = request.form.get('or_label', '')
            session['or_po'] = or_po
            session['or_style'] = or_style
            session['or_org_buyer'] = or_org_buyer
            session['or_color'] = or_color
            session['or_gp_name'] = or_gp_name
            session['or_ext_dely'] = or_ext_dely
            session['or_order_date'] = or_order_date
            session['or_factory'] = or_factory
            session['or_label'] = or_label
            search1 = "%{}%".format(or_po)
            search2 = "%{}%".format(or_style)
            search3 = "%{}%".format(or_org_buyer)
            search4 = "%{}%".format(or_color)
            search5 = "%{}%".format(or_gp_name)
            search6 = "%{}%".format(or_ext_dely)
            search7 = "%{}%".format(or_order_date)
            search8 = "%{}%".format(or_factory)
            search9 = "%{}%".format(or_label)
            all_data = db.session.query(Mocdm_erp.po,Mocdm_pending.label,Mocdm_pending.des,Mocdm_pending.mcn,Mocdm_pending.previous,Mocdm_pending.ext_dely,Mocdm_pending.myanmar,Mocdm_erp.style,Mocdm_erp.buyer_version,Mocdm_erp.pending_buyer,Mocdm_erp.product_name,Mocdm_erp.main_color,Mocdm_erp.season,Mocdm_erp.vessel_date,Mocdm_erp.category,Mocdm_erp.material_classification,Mocdm_erp.material_code,Mocdm_erp.material,Mocdm_erp.material_chinese,Mocdm_erp.size,Mocdm_erp.color,Mocdm_erp.org_consume,Mocdm_erp.unit,Mocdm_erp.loss,Mocdm_erp.consume_point,Mocdm_erp.order_qty,Mocdm_erp.consume,Mocdm_erp.gp,Mocdm_pending.order_date,Mocdm_pending.factory,Mocdm_erp.status).join(Mocdm_pending,(Mocdm_pending.po == Mocdm_erp.po) & (Mocdm_pending.style == Mocdm_erp.style) & (Mocdm_pending.color == Mocdm_erp.main_color) & (Mocdm_pending.org_buyer == Mocdm_erp.pending_buyer),isouter = True).filter((Mocdm_pending.po.like(search1)),(Mocdm_pending.style.like(search2)),(Mocdm_pending.org_buyer.like(search3)),(Mocdm_pending.color.like(search4)),(Mocdm_pending.gp_name.like(search5)),(Mocdm_pending.ext_dely.like(search6)),(Mocdm_pending.order_date.like(search7)),(Mocdm_pending.factory.like(search8)),(Mocdm_pending.label.like(search9))).paginate(per_page=100, page=page_num, error_out=True)
            return render_template("orderlist.html", or_po = or_po,or_style = or_style, org_buyer = or_org_buyer, or_color = or_color, or_gp_name = or_gp_name, or_ext_dely = or_ext_dely, or_order_date = or_order_date, or_factory= or_factory, or_label=or_label, all_data = all_data, orderlist_active="is_active('/orderlist')")
        else:
            all_data = db.session.query(Mocdm_erp.po,Mocdm_pending.label,Mocdm_pending.des,Mocdm_pending.mcn,Mocdm_pending.previous,Mocdm_pending.ext_dely,Mocdm_pending.myanmar,Mocdm_erp.style,Mocdm_erp.buyer_version,Mocdm_erp.pending_buyer,Mocdm_erp.product_name,Mocdm_erp.main_color,Mocdm_erp.season,Mocdm_erp.vessel_date,Mocdm_erp.category,Mocdm_erp.material_classification,Mocdm_erp.material_code,Mocdm_erp.material,Mocdm_erp.material_chinese,Mocdm_erp.size,Mocdm_erp.color,Mocdm_erp.org_consume,Mocdm_erp.unit,Mocdm_erp.loss,Mocdm_erp.consume_point,Mocdm_erp.order_qty,Mocdm_erp.consume,Mocdm_erp.gp,Mocdm_pending.order_date,Mocdm_pending.factory,Mocdm_erp.status).join(Mocdm_pending,(Mocdm_pending.po == Mocdm_erp.po) & (Mocdm_pending.style == Mocdm_erp.style) & (Mocdm_pending.color == Mocdm_erp.main_color) & (Mocdm_pending.org_buyer == Mocdm_erp.pending_buyer),isouter = True).paginate(per_page=100, page=page_num, error_out=True) 
            return render_template('orderlist.html', all_data = all_data, orderlist_active="is_active('/orderlist')")
    except SQLAlchemyError as e:
        current_time = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
        logging.basicConfig(filename= f'error_log.log', level=logging.ERROR)
        logging.error(str(e))
    return redirect(url_for('auth.orderlist'))
    # return f"<td>{all_data.ext_dely}</td>,<td>{all_data.mcn}</td>,<td>{all_data.po}</td>,<td>{all_data.myanmar}</td>,<td>{all_data.ship_to}</td>,<td>{all_data.label}</td>,<td>{all_data.linked_store}</td>,<td>{all_data.des}</td>,<td>{all_data.gp_name}</td>,<td>{all_data.style}</td>,<td>{all_data.org_buyer}</td>,<td>{all_data.color}</td>,<td>{all_data.qty}</td>,<td>{all_data.vessel_date}</td>,<td>{all_data.factory}</td>,<td>{all_data.db_gb_code}</td>,<td>{all_data.sdn_po}</td>,<td>{all_data.customer_po}</td>,<td>{all_data.upc_no}</td>,<td>{all_data.linked_so_no}</td>,<td>{all_data.ref_no}</td>,<td>{all_data.material_log_no}</td>,<td>{all_data.season}</td>,<td>{all_data.buyer_txt}</td>,<td>{all_data.order_date}</td>,<td>{all_data.kmz_id}</td>,<td>{all_data.remark}</td>,<td>{all_data.shpg_job}</td>,<td>{all_data.xfty_date}</td>,<td>{all_data.status}</td>,<td><a href='/pendingUpdate'  class='btn btn-primary' value='{all_data.id}' edit-row-value='{all_data.id}' data-bs-toggle='modal' data-bs-target='#myModal{all_data.id}'>Edit</a></td>"

# all_data = db.session.query(Mocdm_erp.po,Mocdm_pending.label,Mocdm_pending.des,Mocdm_pending.mcn,Mocdm_pending.previous,Mocdm_pending.ext_dely,Mocdm_pending.myanmar,Mocdm_erp.style,Mocdm_erp.buyer_version,Mocdm_erp.pending_buyer,Mocdm_erp.product_name,Mocdm_erp.main_color,Mocdm_erp.season,Mocdm_erp.vessel_date,Mocdm_erp.category,Mocdm_erp.material_classification,Mocdm_erp.material_code,Mocdm_erp.material,Mocdm_erp.material_chinese,Mocdm_erp.size,Mocdm_erp.color,Mocdm_erp.org_consume,Mocdm_erp.unit,Mocdm_erp.loss,Mocdm_erp.consume_point,Mocdm_erp.order_qty,Mocdm_erp.consume,Mocdm_erp.gp,Mocdm_pending.order_date,Mocdm_pending.factory).join(Mocdm_pending,(Mocdm_pending.po == Mocdm_erp.po) & (Mocdm_pending.style == Mocdm_erp.style) & (Mocdm_pending.color == Mocdm_erp.main_color) & (Mocdm_pending.org_buyer == Mocdm_erp.pending_buyer),isouter = True).all()

@auth.route('/download_order', methods=['GET', 'POST'])
def download_order():
    if request.args.get('search') == 'true':
        po = request.args.get('po')
        style = request.args.get('style')
        org_buyer = request.args.get('org_buyer')
        color = request.args.get('color')
        gp_name = request.args.get('gp_name')
        ext_dely = request.args.get('ext_dely')
        order_date = request.args.get('order_date')
        factory = request.args.get('factory')
        label = request.args.get('label')
        po = "%{}%".format(po)
        style = "%{}%".format(style)
        org_buyer = "%{}%".format(org_buyer)
        color = "%{}%".format(color)
        gp_name = "%{}%".format(gp_name)
        ext_dely = "%{}%".format(ext_dely)
        order_date = "%{}%".format(order_date)
        factory = "%{}%".format(factory)
        label = "%{}%".format(label)
        all_data = db.session.query(Mocdm_erp.po,Mocdm_pending.label,Mocdm_pending.des,Mocdm_pending.mcn,Mocdm_pending.previous,Mocdm_pending.ext_dely,Mocdm_pending.myanmar,Mocdm_erp.style,Mocdm_erp.buyer_version,Mocdm_erp.pending_buyer,Mocdm_erp.product_name,Mocdm_erp.main_color,Mocdm_erp.season,Mocdm_erp.vessel_date,Mocdm_erp.category,Mocdm_erp.material_classification,Mocdm_erp.material_code,Mocdm_erp.material,Mocdm_erp.material_chinese,Mocdm_erp.size,Mocdm_erp.color,Mocdm_erp.org_consume,Mocdm_erp.unit,Mocdm_erp.loss,Mocdm_erp.consume_point,Mocdm_erp.order_qty,Mocdm_erp.consume,Mocdm_erp.gp,Mocdm_pending.order_date,Mocdm_pending.factory,Mocdm_erp.status).join(Mocdm_pending,(Mocdm_pending.po == Mocdm_erp.po) & (Mocdm_pending.style == Mocdm_erp.style) & (Mocdm_pending.color == Mocdm_erp.main_color) & (Mocdm_pending.org_buyer == Mocdm_erp.pending_buyer),isouter = True).filter((Mocdm_pending.po.like(po)),(Mocdm_pending.style.like(style)),(Mocdm_pending.org_buyer.like(org_buyer)),(Mocdm_pending.color.like(color)),(Mocdm_pending.gp_name.like(gp_name)),(Mocdm_pending.ext_dely.like(ext_dely)),(Mocdm_pending.order_date.like(order_date)),(Mocdm_pending.factory.like(factory)),(Mocdm_pending.label.like(label))).all()
        wb = Workbook()
        ws = wb.active
        ws.append(['PO','LABEL','DES','D/C','PREVIOUS','DELY','MYANMAR','STYLE','BUYER VERSION','BUYER','PRODUCT NAME','MAIN COLOR','SEASON','VESSEL DATE','CATEGORY','MATERIAL CLASSIFICATION','MATERIAL CODE','MATERIAL NAME IN CHINESE','MATERIAL','SIZE','COLOUR','ORIGINAL CONSUME','UNIT','LOSS','CONSUME POINT','ORDER QTY','CONSUME','GROUP','ORDER DATE','FACTORY','STATUS'])
        for item in all_data:
            ws.append([item.po,item.label,item.des,item.mcn,item.previous.strftime('%m/%d/%Y'),item.ext_dely.strftime('%m/%d/%Y'),item.myanmar,item.style,item.buyer_version,item.pending_buyer,item.product_name,item.main_color,item.season,item.vessel_date.strftime('%m/%d/%Y'),item.category,item.material_classification,item.material_code,item.material,item.material_chinese,item.size,item.color,item.org_consume,item.unit,item.loss,item.consume_point,item.order_qty,item.consume,item.gp,item.order_date.strftime('%m/%d/%Y'),item.factory,item.status])
        file = BytesIO()
        wb.save(file)
        file.seek(0)
        filename = 'combination.xlsx'
        r = Response(
            file.read(),
            headers={
                'Content-Disposition': f'attachment;filename={filename}',
                'Content-Type': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            }
        )
        return r
    else:
        all_data1 = db.session.query(Mocdm_erp.po,Mocdm_pending.label,Mocdm_pending.des,Mocdm_pending.mcn,Mocdm_pending.previous,Mocdm_pending.ext_dely,Mocdm_pending.myanmar,Mocdm_erp.style,Mocdm_erp.buyer_version,Mocdm_erp.pending_buyer,Mocdm_erp.product_name,Mocdm_erp.main_color,Mocdm_erp.season,Mocdm_erp.vessel_date,Mocdm_erp.category,Mocdm_erp.material_classification,Mocdm_erp.material_code,Mocdm_erp.material,Mocdm_erp.material_chinese,Mocdm_erp.size,Mocdm_erp.color,Mocdm_erp.org_consume,Mocdm_erp.unit,Mocdm_erp.loss,Mocdm_erp.consume_point,Mocdm_erp.order_qty,Mocdm_erp.consume,Mocdm_erp.gp,Mocdm_pending.order_date,Mocdm_pending.factory,Mocdm_erp.status).join(Mocdm_pending,(Mocdm_pending.po == Mocdm_erp.po) & (Mocdm_pending.style == Mocdm_erp.style) & (Mocdm_pending.color == Mocdm_erp.main_color) & (Mocdm_pending.org_buyer == Mocdm_erp.pending_buyer),isouter = True).all()
        wb = Workbook()
        ws = wb.active
        ws.append(['PO','LABEL','DES','D/C','PREVIOUS','DELY','MYANMAR','STYLE','BUYER VERSION','BUYER','PRODUCT NAME','MAIN COLOR','SEASON','VESSEL DATE','CATEGORY','MATERIAL CLASSIFICATION','MATERIAL CODE','MATERIAL NAME IN CHINESE','MATERIAL','SIZE','COLOUR','ORIGINAL CONSUME','UNIT','LOSS','CONSUME POINT','ORDER QTY','CONSUME','GROUP','Order Date','FACTORY','STATUS'])
        for item in all_data1:
            ws.append([item.po,item.label,item.des,item.mcn,item.previous.strftime('%m/%d/%Y'),item.ext_dely.strftime('%m/%d/%Y'),item.myanmar,item.style,item.buyer_version,item.pending_buyer,item.product_name,item.main_color,item.season,item.vessel_date.strftime('%m/%d/%Y'),item.category,item.material_classification,item.material_code,item.material,item.material_chinese,item.size,item.color,item.org_consume,item.unit,item.loss,item.consume_point,item.order_qty,item.consume,item.gp,item.order_date.strftime('%m/%d/%Y'),item.factory,item.status])
        file = BytesIO()
        wb.save(file)
        file.seek(0)
        filename = 'combination.xlsx'
        r = Response(
            file.read(),
            headers={
                'Content-Disposition': f'attachment;filename={filename}',
                'Content-Type': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            }
        )
        return r

# @auth.route('/download/orderlist', methods=['GET', 'POST'])
# def download_report():
#     try:
#         all_data = db.session.query(Mocdm_erp.po,Mocdm_pending.label,Mocdm_pending.des,Mocdm_pending.mcn,Mocdm_pending.previous,Mocdm_pending.ext_dely,Mocdm_pending.myanmar,Mocdm_erp.style,Mocdm_erp.buyer_version,Mocdm_erp.pending_buyer,Mocdm_erp.product_name,Mocdm_erp.main_color,Mocdm_erp.season,Mocdm_erp.vessel_date,Mocdm_erp.category,Mocdm_erp.material_classification,Mocdm_erp.material_code,Mocdm_erp.material,Mocdm_erp.material_chinese,Mocdm_erp.size,Mocdm_erp.color,Mocdm_erp.org_consume,Mocdm_erp.unit,Mocdm_erp.loss,Mocdm_erp.consume_point,Mocdm_erp.order_qty,Mocdm_erp.consume,Mocdm_erp.gp,Mocdm_pending.order_date,Mocdm_pending.factory,Mocdm_erp.status).join(Mocdm_pending,(Mocdm_pending.po == Mocdm_erp.po) & (Mocdm_pending.style == Mocdm_erp.style) & (Mocdm_pending.color == Mocdm_erp.main_color) & (Mocdm_pending.org_buyer == Mocdm_erp.pending_buyer),isouter = True).all()
#         wb = Workbook()
#         ws = wb.active
#         ws.append(['PO#','LABEL','DES','D/C','PREVIOUS','DELY','MYANMAR','STYLE','BUYER VERSION','BUYER','PRODUCT NAME','MAIN COLOR','SEASON','VESSEL DATE','CATEGORY','MATERIAL CLASSIFICATION','MATERIAL CODE','MATERIAL NAME IN CHINESE','MATERIAL','SIZE','COLOUR','ORIGINAL CONSUME','UNIT','LOSS','CONSUME POINT','ORDER QTY','CONSUME','GROUP','ORDER DATE','FACTORY','STATUS'])
#         for item in all_data:
#             ws.append([item.po,item.label,item.des,item.mcn,item.previous.strftime('%d/%m/%Y'),item.ext_dely.strftime('%d/%m/%Y'),item.myanmar,item.style,item.buyer_version,item.pending_buyer,item.product_name,item.main_color,item.season,item.vessel_date.strftime('%d/%m/%Y'),item.category,item.material,item.material_classification,item.material_code,item.material,item.material_chinese,item.size,item.color,item.org_consume,item.unit,item.loss,item.consume_point,item.order_qty,item.consume,item.gp,item.order_date,item.factory,item.status])
#         file = BytesIO()
#         wb.save(file)
#         file.seek(0)
#         filename = 'combination.xlsx'
#         response = Response(
#             file.read(),
#             headers={
#                 'Content-Disposition': f'attachment;filename={filename}',
#                 'Content-Type': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
#             }
#         )
#         return response
#     except SQLAlchemyError as e:
#         current_time = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
#         logging.basicConfig(filename= f'error_log.log', level=logging.ERROR)
#         logging.error(str(e))
#     return redirect(url_for('auth.orderlist'))

@auth.route('/consumptionreport' , defaults={'page_num':1})
@auth.route('/consumptionreport/<int:page_num>', methods=['GET','POST'])
@login_required
def consumptionreport(page_num):
    if request.method=='GET':
        factory = request.args.get('factory')
        gp_name = request.args.get('gp_name')
        des = request.args.get('des')
        qty = request.args.get('qty')
        ext_dely = request.args.get('ext_dely')
        date_object = datetime.strptime(ext_dely, "%Y-%m-%d").date()
        new_date_string = date_object.strftime("%m/%d/%Y")
        style = request.args.get('style')
        org_buyer = request.args.get('org_buyer')
        all_data = db.session.query(Mocdm_erp.id.label("erpid"),Mocdm_consumption.id,Mocdm_erp.category,Mocdm_erp.material,Mocdm_erp.color,Mocdm_erp.consume_point,Mocdm_erp.consume,Mocdm_consumption.issued_qty,Mocdm_consumption.balance,Mocdm_consumption.date,Mocdm_consumption.issued_by_leader,Mocdm_consumption.factory_line,Mocdm_consumption.reciever,Mocdm_consumption.remark,Mocdm_pending.qty).join(Mocdm_pending,(Mocdm_pending.po == Mocdm_erp.po) & (Mocdm_pending.color == Mocdm_erp.main_color) & (Mocdm_pending.style == Mocdm_erp.style) & (Mocdm_pending.org_buyer == Mocdm_erp.pending_buyer),isouter = True).join(Mocdm_consumption,(Mocdm_consumption.erp_id == Mocdm_erp.id),isouter = True).filter(Mocdm_pending.factory == factory,Mocdm_pending.gp_name == gp_name,Mocdm_pending.des == des,Mocdm_pending.ext_dely == ext_dely,Mocdm_pending.style == style,Mocdm_pending.org_buyer == org_buyer).paginate(per_page=100, page=page_num, error_out=True)
        return render_template('consumptionlistreport.html',factory=factory, gp_name=gp_name, qty=qty, des=des, ext_dely=new_date_string,ext_delys=ext_dely, style=style, org_buyer=org_buyer, all_data = all_data)


@auth.route('/searchcreport', methods=['GET','POST'])
@login_required
def searchcreport():
    try:
        if request.method=='POST':
            ext_dely = request.form['ext_dely']
            qty = request.form['qty']
            gp_name = request.form['gp_name']
            style = request.form['style']
            search1 = "%{}%".format(ext_dely)
            search2 = "%{}%".format(qty)
            search3 = "%{}%".format(gp_name)
            search4 = "%{}%".format(style)
            all_data = Mocdm_pending.query.filter((Mocdm_pending.ext_dely.like(search1)),(Mocdm_pending.qty.like(search2)),(Mocdm_pending.gp_name.like(search3)),(Mocdm_pending.style.like(search4))).all()
            if not all_data:
                return "no data"
            else:
                return render_template("searchpending.html", all_data = all_data)   
    except SQLAlchemyError as e:
        current_time = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
        logging.basicConfig(filename= f'error_log.log', level=logging.ERROR)
        logging.error(str(e))
    return redirect(url_for('auth.profile'))
       

@auth.route('/consumptionreportUpdate', methods=['GET', 'POST'])
@login_required
def consumptionreportUpdate():
    if request.method=='POST':
        data = Mocdm_consumption.query.get(request.form.get('id'))
        if data:
            style = request.args.get('style')
            group_name = request.args.get('group_name')
            qty_no = request.args.get('qty_no')
            dely_date = request.args.get('dely_date')
            buyer = request.args.get('buyer')
            all_data = Mocdm_consumption(id=request.form['id'],issued_qty=request.form['issued_qty'] or 0,balance=request.form['balance'] or 0,date=request.form['date'] or None,issued_by_leader=request.form['issued_by_leader'] or None,factory_line=request.form['factory_line'] or None,reciever=request.form['reciever'] or None,remark=request.form['remark'] or None)
            db.session.merge(all_data)
            db.session.commit()
            return redirect(request.referrer or url_for('auth.consumptionreport') + '?style=style&group_name=group_name&qty_no=qty_no&dely_date=dely_date&buyer=buyer')
        else:
            style = request.args.get('style')
            group_name = request.args.get('group_name')
            qty_no = request.args.get('qty_no')
            dely_date = request.args.get('dely_date')
            buyer = request.args.get('buyer')
            all_data = Mocdm_consumption(erp_id=request.form['erpid'] ,issued_qty=request.form['issued_qty'] or 0,balance=request.form['balance'] or 0,date=request.form['date'] or None,issued_by_leader=request.form['issued_by_leader'] or None,factory_line=request.form['factory_line'] or None,reciever=request.form['reciever'] or None,remark=request.form['remark'] or None)
            db.session.add(all_data)
            db.session.commit()
            return redirect(request.referrer or url_for('auth.consumptionreport') + '?style=style&group_name=group_name&qty_no=qty_no&dely_date=dely_date&buyer=buyer')

@auth.route('/download/consumptionreport', methods=['GET', 'POST'])
def download_consumptionreportreport():
    factory = request.args.get('factory')
    gp_name = request.args.get('gp_name')
    qty = request.args.get('qty')
    ext_dely = request.args.get('ext_dely')
    date_object = datetime.strptime(ext_dely, "%m/%d/%Y").date()
    new_date_string = date_object.strftime("%Y-%m-%d")
    style = request.args.get('style')
    org_buyer = request.args.get('org_buyer')
    all_data = db.session.query(Mocdm_erp.id.label("erpid"),Mocdm_consumption.id,Mocdm_erp.category,Mocdm_erp.material,Mocdm_erp.color,Mocdm_erp.consume_point,Mocdm_erp.consume,Mocdm_erp.order_qty,Mocdm_consumption.issued_qty,Mocdm_consumption.balance,Mocdm_consumption.date,Mocdm_consumption.issued_by_leader,Mocdm_consumption.factory_line,Mocdm_consumption.reciever,Mocdm_consumption.remark,Mocdm_pending.qty).join(Mocdm_pending,(Mocdm_pending.po == Mocdm_erp.po) & (Mocdm_pending.color == Mocdm_erp.main_color) & (Mocdm_pending.style == Mocdm_erp.style) & (Mocdm_pending.org_buyer == Mocdm_erp.pending_buyer),isouter = True).join(Mocdm_consumption,(Mocdm_consumption.erp_id == Mocdm_erp.id),isouter = True).filter(Mocdm_pending.factory == factory,Mocdm_pending.gp_name == gp_name,Mocdm_pending.ext_dely == new_date_string,Mocdm_pending.style == style,Mocdm_pending.org_buyer == org_buyer).all()       
    wb = Workbook()
    ws = wb.active
    ws.merge_cells('J1:M1')
    ws['J1'] = 'Manager'
    ws['J2'] = ''
    ws['J1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.append(['Factory','Group Name','Qty','Dely Date','Style','Buyer'])
    ws.append([factory,group_name,qty_no,ext_dely,style,buyer])
    ws.append(['CATEGORY','MATERIAL','COLOUR','QTY','CONSUME','TOTAL QTY','ISSUED QTY','BALANCE','DATE','ISSUED BY (Leader)','Factory line','RECEIVER','REMARK','SIGN'])
    for item in all_data:
        ws.append([item.category,item.material,item.color,item.qty,item.consume_point,item.consume,item.issued_qty,item.balance,item.date,item.issued_by_leader,item.factory_line,item.reciever,item.remark])
    file = BytesIO()
    wb.save(file)
    file.seek(0)
    filename = 'consumptionsreport.xlsx'
    r = Response(
        file.read(),
        headers={
            'Content-Disposition': f'attachment;filename={filename}',
            'Content-Type': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        }
    )
    return r

@auth.route("/scheduleUpload", methods=['POST'])
@login_required
def scheduleUpload():
    try:
        if request.method == 'POST':
            excel_file = request.files['file']
            col_names = ['LINE','DELY','QTY','Target(H/W)','Balance(PVC)', 'Zip & Thread' , 'Group','Style','Version','Buyer','Factory','DES','Total','Data Date']
            df = pd.read_excel(excel_file,names=col_names,header = None,skiprows=1)
            df = df.fillna('')
            for i,row in df.iterrows():
                c = row['Style']
                d = c.replace(" ", "")
                all_data = Mocdm_schedule(line=row['LINE'],dely=row['DELY'],qty=row['QTY'],target=row['Target(H/W)'],balance=row['Balance(PVC)'],zip_thread=row['Zip & Thread'],gp=row['Group'],style=d,version=row['Version'],buyer=row['Buyer'],factory=row['Factory'],des=row['DES'],total=row['Total'],data_date=row['Data Date'])
                db.session.add(all_data)
                db.session.commit()
            return redirect(url_for('auth.schedulelist'))
    except SQLAlchemyError as e:
        current_time = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
        logging.basicConfig(filename= f'error_log.log', level=logging.ERROR)
        logging.error(str(e))
    return redirect(url_for('auth.schedulelist'))
        

@auth.route('/schedulelist', defaults={'page_num':1})
@auth.route('/schedulelist/<int:page_num>', methods=['GET', 'POST'])
@login_required
def schedulelist(page_num):
    if request.method=='GET':
        search_factory = request.args.get("search_factory")
        if search_factory:
            all_data = Mocdm_schedule.query.filter(Mocdm_schedule.factory == search_factory).all()
        else:
            all_data = Mocdm_schedule.query.all()
        # Send default values when db rows are empty
        if len(all_data) <= 0:
            return render_template('schedulereport.html',
                               all_data = all_data,
                               today = datetime.now().strftime("%m/%d/%Y"),
                               index_range = 0,
                               dates = [],
                               occurrence = 0,
                               search_factory = search_factory,
                               schedule_active="is_active('/schedule')")

        # Get start and end date from db
        start_date = db.session.query(func.min(Mocdm_schedule.data_date)).scalar()
        end_date = db.session.query(func.max(Mocdm_schedule.data_date)).scalar()
        
        # Get the list of dates between start and end
        get_all_date = pd.date_range(start_date, end_date, freq='d')
        date_range = [dict(days = days) for days in get_all_date]

        # Mocdm_schedule.query.paginate(per_page=100, page=page_num, error_out=True)
        all_data = [item.__dict__ for item in all_data]

        # # Find all occurrences
        # occurrence = collections.Counter()
        # for items in all_data:
        #     occurrence[items["data_date"]] +=1
        # occurrence_count = occurrence.most_common(1)[0][1]
        line_list = [i['line'] for i in all_data]

        result = []
        for row in range(len(line_list)):
            splice_by_row = []
            column_range = [dict(days = days) for days in get_all_date]

            # Get data by each row
            for date_index in range(len(column_range)):
                for column in all_data:
                    if column["data_date"] == date_range[date_index]["days"].date() and column['line'] == line_list[row]:
                        splice_by_row.append({
                            "idx": date_index,
                            "data": column
                        })
                        all_data.remove(column)
                        break
                    else:
                        pass
            
            # Make data by its idx. 
            for i in range(len(column_range)):
                for idx, x in enumerate([a['idx'] for a in splice_by_row]):
                    if x == i:
                        column_range[i]['existed'] = True
                        column_range[i]['data_idx'] = idx
                    else:
                        pass
            result.append({
                "list_of_placement": column_range,
                "data": splice_by_row,
            })

        return render_template('schedulereport.html',
                               all_data = result,
                               today = datetime.now().strftime("%m/%d/%Y"),
                               index_range = len(date_range),
                               dates = date_range,
                               line_list = line_list,
                               occurrence = len(line_list),
                               search_factory = search_factory,
                               schedule_active="is_active('/schedule')")

@auth.route('/schedule/update/<id>', methods = ['GET', 'POST'])
def updateSchedule(id):
    all_data = Mocdm_schedule.query.get(id)
    if request.method == 'GET':
        return render_template("scheduleupdate.html", all_data=all_data)
    else:
        all_data.dely_date = request.form['dely_date']
        all_data.qty = request.form['qty']
        all_data.target = request.form['target']
        all_data.balance = request.form['balance']
        all_data.zip_thread = request.form['zip']
        all_data.group = request.form['group']
        all_data.version = request.form['version']
        all_data.style = request.form['style']
        all_data.data_date = request.form['data_date']
        all_data.total = request.form['total']
        all_data.factory = request.form['factory']
        if "image_path" in request.files and request.files['image_path'].filename:
            file = request.files['image_path']
            filename = file.filename
            file_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
            file.save(file_path)
            os.remove(all_data.image_path)
            all_data.image_path = file_path
        db.session.commit()
        return redirect(url_for("auth.schedulelist"))

@auth.route('/deleteSchedule/<id>/', methods = ['POST'])
def deleteSchedule(id):
    all_data = Mocdm_schedule.query.get(id)
    db.session.delete(all_data)
    db.session.commit()
    return redirect(url_for('auth.schedulelist'))

@auth.route('/image/<int:id>')
def view_image(id):
    image = Mocdm_schedule.query.filter_by(id=id).first()
    return render_template('schedulereport.html', image=image)


@auth.route('/ddCon', defaults={'page_num':1}, methods=['GET','POST'])
@auth.route('/ddCon/<int:page_num>', methods=['GET','POST'])
@login_required
def ddCon(page_num):
    current_time = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    logging.basicConfig(filename= f'error_log.log', level=logging.DEBUG)
    logging.error('reach ddcon')
    if request.method=='POST' or request.args.get('search')=='True' :
        logging.error('reach post')
        if (request.args.get('con_po')): 
            con_po = request.args.get('con_po')
        else:
            con_po = request.form.get('con_po', '')
        if (request.args.get('con_style')): 
            con_style = request.args.get('con_style')
        else:
            con_style = request.form.get('con_style', '')

        if (request.args.get('con_org_buyer')): 
            con_org_buyer = request.args.get('con_org_buyer')
        else:
            con_org_buyer = request.form.get('con_org_buyer', '')

        if (request.args.get('con_color')): 
            con_color = request.args.get('con_color')
        else:
            con_color = request.form.get('con_color', '')

        if (request.args.get('con_gp_name')): 
            con_gp_name = request.args.get('con_gp_name')
        else:
            con_gp_name = request.form.get('con_gp_name', '')

        if (request.args.get('con_ext_dely')): 
            con_ext_dely = request.args.get('con_ext_dely')
        else:
            con_ext_dely = request.form.get('con_ext_dely', '')

        if (request.args.get('con_order_date')): 
            con_order_date = request.args.get('con_order_date')
        else:
            con_order_date = request.form.get('con_order_date', '')

        if (request.args.get('con_factory')): 
            con_factory = request.args.get('con_factory')
        else:
            con_factory = request.form.get('con_factory', '')

        if (request.args.get('con_des')): 
            con_des = request.args.get('con_des')
        else:
            con_des = request.form.get('con_des', '')
        session['con_po'] = con_po
        session['con_style'] = con_style
        session['con_org_buyer'] = con_org_buyer
        session['con_color'] = con_color
        session['con_gp_name'] = con_gp_name
        session['con_ext_dely'] = con_ext_dely
        session['con_order_date'] = con_order_date
        session['con_factory'] = con_factory
        session['con_des'] = con_des
        search1 = "%{}%".format(con_po)
        search2 = "%{}%".format(con_style)
        search3 = "%{}%".format(con_org_buyer)
        search4 = "%{}%".format(con_color)
        search5 = "%{}%".format(con_gp_name)
        search6 = "%{}%".format(con_ext_dely)
        search7 = "%{}%".format(con_order_date)
        search8 = "%{}%".format(con_factory)
        search9 = "%{}%".format(con_des)
        all_data = db.session.query(func.sum(Mocdm_pending.qty).label('qty'),Mocdm_pending.factory, Mocdm_pending.gp_name,Mocdm_pending.ext_dely, Mocdm_pending.style, Mocdm_pending.org_buyer).filter((Mocdm_pending.po.like(search1)),(Mocdm_pending.style.like(search2)),(Mocdm_pending.org_buyer.like(search3)),(Mocdm_pending.color.like(search4)),(Mocdm_pending.gp_name.like(search5)),(Mocdm_pending.ext_dely.like(search6)),(Mocdm_pending.order_date.like(search7)),(Mocdm_pending.factory.like(search8)),(Mocdm_pending.des.like(search9))).group_by(Mocdm_pending.factory, Mocdm_pending.gp_name,Mocdm_pending.des,Mocdm_pending.ext_dely, Mocdm_pending.style, Mocdm_pending.org_buyer).paginate(per_page=100, page=page_num, error_out=True)
        return render_template("consumpList.html", con_po=con_po,con_style=con_style,con_org_buyer=con_org_buyer,con_color=con_color,con_order_date=con_order_date,con_des=con_des,con_gp_name=con_gp_name,con_ext_dely=con_ext_dely,con_factory=con_factory,all_data=all_data, dd_active="is_active('/dd')")
    else:
        logging.error('reach not post')
        logging.error(request.method)
        all_data = db.session.query(func.sum(Mocdm_pending.qty).label('qty'),Mocdm_pending.factory, Mocdm_pending.gp_name,Mocdm_pending.ext_dely, Mocdm_pending.style, Mocdm_pending.org_buyer,Mocdm_pending.des).group_by(Mocdm_pending.factory, Mocdm_pending.gp_name,Mocdm_pending.ext_dely, Mocdm_pending.style, Mocdm_pending.org_buyer,Mocdm_pending.des).paginate(per_page=100, page=page_num, error_out=True)
        return render_template('consumpList.html', all_data=all_data, dd_active="is_active('/dd')")


# @auth.route('/ddd', defaults={'page_num':1}, methods=['GET', 'POST'])
# @auth.route('/ddd/<int:page_num>', methods=['GET', 'POST'])
# @login_required
# def ddd(page_num):
#     if request.method=='POST' or request.args.get('search')=='True' :
#         if (request.args.get('po')): 
#             po = request.args.get('po')
#         else:
#             po = request.form.get('po', '')

#         if (request.args.get('style')): 
#             style = request.args.get('style')
#         else:
#             style = request.form.get('style', '')

#         if (request.args.get('org_buyer')): 
#             org_buyer = request.args.get('org_buyer')
#         else:
#             org_buyer = request.form.get('org_buyer', '')

#         if (request.args.get('color')): 
#             color = request.args.get('color')
#         else:
#             color = request.form.get('color', '')

#         if (request.args.get('gp_name')): 
#             gp_name = request.args.get('gp_name')
#         else:
#             gp_name = request.form.get('gp_name', '')

#         if (request.args.get('ext_dely')): 
#             ext_dely = request.args.get('ext_dely')
#         else:
#             ext_dely = request.form.get('ext_dely', '')

#         if (request.args.get('order_date')): 
#             order_date = request.args.get('order_date')
#         else:
#             order_date = request.form.get('order_date', '')

#         if (request.args.get('factory')): 
#             factory = request.args.get('factory')
#         else:
#             factory = request.form.get('factory', '')

#         if (request.args.get('des')): 
#             des = request.args.get('des')
#         else:
#             des = request.form.get('des', '')
#         session['po'] = po
#         session['style'] = style
#         session['org_buyer'] = org_buyer
#         session['color'] = color
#         session['gp_name'] = gp_name
#         session['ext_dely'] = ext_dely
#         session['order_date'] = order_date
#         session['factory'] = factory
#         session['des'] = des
#         search1 = "%{}%".format(po)
#         search2 = "%{}%".format(style)
#         search3 = "%{}%".format(org_buyer)
#         search4 = "%{}%".format(color)
#         search5 = "%{}%".format(order_date)
#         search6 = "%{}%".format(des)
#         search7 = "%{}%".format(gp_name)
#         search8 = "%{}%".format(ext_dely)
#         search9 = "%{}%".format(factory)
#         all_data = Mocdm_pending.query.filter((Mocdm_pending.po.like(search1)),(Mocdm_pending.style.like(search2)),(Mocdm_pending.org_buyer.like(search3)),(Mocdm_pending.color.like(search4)),(Mocdm_pending.order_date.like(search5)),(Mocdm_pending.des.like(search6)),(Mocdm_pending.gp_name.like(search7)),(Mocdm_pending.ext_dely.like(search8)),(Mocdm_pending.factory.like(search9))).group_by(Mocdm_pending.factory, Mocdm_pending.gp_name, Mocdm_pending.qty,Mocdm_pending.ext_dely, Mocdm_pending.style, Mocdm_pending.org_buyer).paginate(per_page=100, page=page_num, error_out=True)
#         return render_template("consumpList.html", po=po,style=style,org_buyer=org_buyer,color=color,order_date=order_date,des=des,gp_name=gp_name,ext_dely=ext_dely,factory=factory,all_data=all_data, consump_active="is_active('/consump')")
#     else:
#         all_data = db.session.query(Mocdm_pending.factory, Mocdm_pending.gp_name, Mocdm_pending.qty, 
#                         Mocdm_pending.ext_dely, Mocdm_pending.style, Mocdm_pending.org_buyer)\
#                 .group_by(Mocdm_pending.factory, Mocdm_pending.gp_name, Mocdm_pending.qty, 
#                             Mocdm_pending.ext_dely, Mocdm_pending.style, Mocdm_pending.org_buyer).paginate(per_page=100, page=page_num, error_out=True)
#         return render_template('consumpList.html', all_data=all_data, consump_active="is_active('/consump')")
    

# @auth.route('/consump_search', methods=['GET', 'POST'])
# def consump_search():
#     if request.method=='POST':
#         po = request.form['po']
#         style = request.form['style']
#         org_buyer = request.form['org_buyer']
#         color = request.form['color']
#         order_date = request.form['order_date']
#         des = request.form['des']
#         gp_name = request.form['gp_name']
#         ext_dely = request.form['ext_dely']
#         factory = request.form['factory']
#         session['po'] = po
#         session['style'] = style
#         session['org_buyer'] = org_buyer
#         session['color'] = color
#         session['gp_name'] = gp_name
#         session['ext_dely'] = ext_dely
#         session['order_date'] = order_date
#         session['factory'] = factory
#         session['des'] = des
#         search1 = "%{}%".format(po)
#         search2 = "%{}%".format(style)
#         search3 = "%{}%".format(org_buyer)
#         search4 = "%{}%".format(color)
#         search5 = "%{}%".format(order_date)
#         search6 = "%{}%".format(des)
#         search7 = "%{}%".format(gp_name)
#         search8 = "%{}%".format(ext_dely)
#         search9 = "%{}%".format(factory)
#         all_data = Mocdm_pending.query.filter((Mocdm_pending.po.like(search1)),(Mocdm_pending.style.like(search2)),(Mocdm_pending.org_buyer.like(search3)),(Mocdm_pending.color.like(search4)),(Mocdm_pending.order_date.like(search5)),(Mocdm_pending.des.like(search6)),(Mocdm_pending.gp_name.like(search7)),(Mocdm_pending.ext_dely.like(search8)),(Mocdm_pending.factory.like(search9))).group_by(Mocdm_pending.factory, Mocdm_pending.gp_name, Mocdm_pending.qty, 
#                         Mocdm_pending.ext_dely, Mocdm_pending.style, Mocdm_pending.org_buyer).all()
#         return render_template("consumpList.html", all_data = all_data, consump_active="is_active('/consump')")

@auth.route('/consumption_list_report' , defaults={'page_num':1})
@auth.route('/consumption_list_report/<int:page_num>', methods=['GET','POST'])
@login_required
def consumption_list_report(page_num):
    try:
        if request.method=='GET':
            factory = request.args.get('factory')
            gp_name = request.args.get('gp_name')
            des = request.args.get('des')
            qty = request.args.get('qty')
            ext_dely = request.args.get('ext_dely')
            date_object = datetime.strptime(ext_dely, "%Y-%m-%d").date()
            new_date_string = date_object.strftime("%m/%d/%Y")
            style = request.args.get('style')
            org_buyer = request.args.get('org_buyer')
            all_data = db.session.query(Mocdm_erp.id.label("erpid"),Mocdm_consumption.id,Mocdm_erp.category,Mocdm_erp.material,Mocdm_erp.color,Mocdm_erp.consume_point,Mocdm_erp.consume,Mocdm_consumption.issued_qty,Mocdm_consumption.balance,Mocdm_consumption.date,Mocdm_consumption.issued_by_leader,Mocdm_consumption.factory_line,Mocdm_consumption.reciever,Mocdm_consumption.remark,Mocdm_pending.qty).join(Mocdm_pending,(Mocdm_pending.po == Mocdm_erp.po) & (Mocdm_pending.color == Mocdm_erp.main_color) & (Mocdm_pending.style == Mocdm_erp.style) & (Mocdm_pending.org_buyer == Mocdm_erp.pending_buyer),isouter = True).join(Mocdm_consumption,(Mocdm_consumption.erp_id == Mocdm_erp.id),isouter = True).filter(Mocdm_pending.factory == factory,Mocdm_pending.gp_name == gp_name,Mocdm_pending.des == des,Mocdm_pending.ext_dely == ext_dely,Mocdm_pending.style == style,Mocdm_pending.org_buyer == org_buyer).paginate(per_page=100, page=page_num, error_out=True)
            return render_template('consumptionlistreport.html',factory=factory, gp_name=gp_name, qty=qty, des=des, ext_dely=new_date_string,ext_delys=ext_dely, style=style, org_buyer=org_buyer, all_data = all_data)
    except SQLAlchemyError as e:
        current_time = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
        logging.basicConfig(filename= f'error_log.log', level=logging.DEBUG)
        logging.error(str(e))
    return render_template('consumptionlistreport.html',factory=factory, gp_name=gp_name, qty=qty,des=des,ext_delys=ext_dely, ext_dely=new_date_string, style=style, org_buyer=org_buyer, all_data = all_data)

@auth.route('/download/consumption_list_report', methods=['GET', 'POST'])
def download_consumptionlistreportreport():
        factory = request.args.get('factory')
        gp_name = request.args.get('gp_name')
        qty = request.args.get('qty')
        ext_dely = request.args.get('ext_dely')
        date_object = datetime.strptime(ext_dely, "%m/%d/%Y").date()
        new_date_string = date_object.strftime("%Y-%m-%d")
        style = request.args.get('style')
        org_buyer = request.args.get('org_buyer')
        all_data = db.session.query(Mocdm_erp.id.label("erpid"),Mocdm_consumption.id,Mocdm_erp.category,Mocdm_erp.material,Mocdm_erp.color,Mocdm_erp.consume_point,Mocdm_erp.consume,Mocdm_erp.order_qty,Mocdm_pending.qty,Mocdm_consumption.issued_qty,Mocdm_consumption.balance,Mocdm_consumption.date,Mocdm_consumption.issued_by_leader,Mocdm_consumption.factory_line,Mocdm_consumption.reciever,Mocdm_consumption.remark).join(Mocdm_pending,(Mocdm_pending.po == Mocdm_erp.po) & (Mocdm_pending.color == Mocdm_erp.main_color) & (Mocdm_pending.style == Mocdm_erp.style) & (Mocdm_pending.org_buyer == Mocdm_erp.pending_buyer),isouter = True).join(Mocdm_consumption,(Mocdm_consumption.erp_id == Mocdm_erp.id),isouter = True).filter(Mocdm_pending.factory == factory,Mocdm_pending.gp_name == gp_name,Mocdm_pending.ext_dely == new_date_string,Mocdm_pending.style == style,Mocdm_pending.org_buyer == org_buyer).all()       
        wb = Workbook()
        ws = wb.active
        ws.merge_cells('J1:M1')
        ws['J1'] = 'Manager'
        ws['J2'] = ''
        ws['J1'].alignment = Alignment(horizontal='center', vertical='center')
        ws.append(['Factory','Group','Qty','Dely','STYLE','BUYER'])
        ws.append([factory,gp_name,qty,ext_dely,style,org_buyer])
        ws.append(['CATEGORY','MATERIAL','COLOUR','QTY','CONSUME','TOTAL QTY','ISSUED QTY','BALANCE','DATE','ISSUED BY (Leader)','Factory line','RECEIVER','REMARK','SIGN'])
        for item in all_data:
            ws.append([item.category,item.material,item.color,item.qty,item.consume_point,item.consume,item.issued_qty,item.balance,item.date,item.issued_by_leader,item.factory_line,item.reciever,item.remark])
        file = BytesIO()
        wb.save(file)
        file.seek(0)
        filename = 'consumptionsreport.xlsx'
        r = Response(
            file.read(),
            headers={
                'Content-Disposition': f'attachment;filename={filename}',
                'Content-Type': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            }
        )
        return r


@auth.route('/deltamate', methods=['GET', 'POST'])
def deltamate():
    data_date = []
    data_arr = {}
    final_arr = {}
    response = {}

    s_date = datetime.strptime('2023-03-13', '%Y-%m-%d')
    e_date = datetime.strptime('2023-03-28', '%Y-%m-%d')

    while s_date <= e_date:
        data_date.append(s_date.strftime('%Y-%m-%d'))
        s_date += timedelta(days=1)

    for date in data_date:
        final_arr[date] = {}

    result = [
        {'line':1, 'del_date':'2023-05-05', 'qty':700, 'target':1000, 'balance':1000, 'zip_thread':1000, 'group':'MORA', 'style':'LJ-0423C', 'version':'2', 'buyer':'AF3114E0305', 'data_date':'2023-03-13'},
        {'line':1, 'del_date':'2023-05-05', 'qty':700, 'target':1000, 'balance':1000, 'zip_thread':1000, 'group':'MORA', 'style':'LJ-0423C', 'version':'2', 'buyer':'AF3114E0305', 'data_date':'2023-03-14'},
        {'line':1, 'del_date':'2023-05-05', 'qty':700, 'target':1000, 'balance':1000, 'zip_thread':1000, 'group':'MORA', 'style':'LJ-0423C', 'version':'2', 'buyer':'AF3114E0305', 'data_date':'2023-03-15'},
        {'line':2, 'del_date':'2023-05-05', 'qty':700, 'target':1000, 'balance':1000, 'zip_thread':1000, 'group':'MORA', 'style':'LJ-0423C', 'version':'2', 'buyer':'AF3114E0305', 'data_date':'2023-03-13'},
        {'line':3, 'del_date':'2023-05-05', 'qty':700, 'target':1000, 'balance':1000, 'zip_thread':1000, 'group':'MORA', 'style':'LJ-0423C', 'version':'2', 'buyer':'AF3114E0305', 'data_date':'2023-03-13'},
        {'line':3, 'del_date':'2023-05-05', 'qty':700, 'target':1000, 'balance':1000, 'zip_thread':1000, 'group':'MORA', 'style':'LJ-0423C', 'version':'2', 'buyer':'AF3114E0305', 'data_date':'2023-03-15'}
    ]

    for row in result:
        data_arr.setdefault(row['line'], []).append(row)

    for key, row in data_arr.items():
        for x in data_date:
            final_arr[x] = {}

        for row in row:
            final_arr[row['data_date']] = row

        response[key] = final_arr

    return render_template('dd.html', data_date=data_date)